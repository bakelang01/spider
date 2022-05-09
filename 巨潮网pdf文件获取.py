# coding = utf-8

import requests
import openpyxl as ex
import os

def get_data(key,page_number,date="2022-01-01~2022-05-07",start=1,year=False,colum='szse'):
    '''

    :param key: 关键词搜索
    :param page_number: 获取页数末页
    :param date: 指定某时间段
    :param start: 开始爬取的页数
    :param year: 是否获取年报（其实可以换成其他的筛选条件）
    :param colum: 不同的板块
    :return: 列表中字典数据
    '''
    totle_new=[]
    url = 'http://www.cninfo.com.cn/new/hisAnnouncement/query'
    header = {
        "Accept-Encoding": "gzip,deflate",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Connection": "keep-alive",
        # "Content-Length":" 163",
        "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
        # "Cookie":" JSESSIONID=08B67A430D70E3FB661A8B326AE32F53; _sp_ses.2141=*; insert_cookie=45380249; routeId=.uc1; SID=e394bfac-1e5d-4af1-ab53-79c3fb492a6b; _sp_id.2141=4db067f8-9e72-4b81-96fa-3d1391b2f2a8.1651199342.6.1651845467.1651835944.c75c87e2-e9ed-498f-8efe-42ded97d2de5",
        "Host": "www.cninfo.com.cn",
        "Origin": "http://www.cninfo.com.cn",
        "Referer": "http://www.cninfo.com.cn/new/commonUrl/pageOfSearch?url=disclosure/list/search&lastPage=index",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.54 Safari/537.36",
        "X-Requested-With": "XMLHttpRequest",
    }
    if year :
        year_1 = "category_ndbg_szsh"
    else:
        year_1 = ""
    for i in range(start,page_number):
        print(f'第{i}页数据获取中')
        data = {
            "pageNum": i,
            "pageSize": "30",
            "column": colum,
            "tabName": "fulltext",
            "plate": "",
            "stock": "",
            "searchkey": key,
            "secid": "",
            "category": year_1,
            "trade": "",
            "seDate": date,
            "sortName": "",
            "sortType": "",
            "isHLtitle": "true",
        }
        datas = requests.post(url=url, headers=header, data=data, timeout=5)
        if datas.status_code == 200:
            datas.encoding = 'utf-8'
            pdf_li = datas.json()["announcements"]
            for li in pdf_li:
                di = {}
                di['code'] = li["secCode"]
                di['name'] = li['secName']
                di['title'] = li["announcementTitle"]
                di["announcementTime"] = li["announcementTime"]
                di["pdfurl"] = li['adjunctUrl']
                di["date"] = li['adjunctUrl'].split('/')[1]
                di["orgId"] = li["orgId"]
                di["announcementId"] = li["announcementId"]
                totle_new.append(di)
    # with open('datas.txt', 'w', encoding='utf-8') as f:
    #     f.write(str(totle_new))
    #     print('保存成功！')
    return totle_new

def fenxidata(data):
    '''

    :param data: 传入待分析数据
    :return: 返回筛选后数据
    '''
    useful=[]
    codes=[]
    for di in data:
        di['title'] = ''.join(di['title'].split('<em>ESG</em>'))
        di['title'] = ''.join(di['title'].split('（）'))
        di['name'] = ''.join(di['name'].split('*'))
        if '英文' in di['title'] or '摘要' in di['title'] :
            continue
        elif di["code"]  in codes:
            continue
        else:
            codes.append(di["code"])
            useful.append(di)
    return useful

def down_pdf(pdf_li):
    '''

    :param pdf_li: 进行数据下载
    :return:无返回
    '''
    if not os.path.exists('data') :
        os.mkdir('data')
    for pdf in pdf_li:
        url = 'https://static.cninfo.com.cn/'+pdf['pdfurl']
        hesder={
            'User-Agent' :'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.41 Safari/537.36 Edg/101.0.1210.32'
        }
        response = requests.get(url=url,headers=hesder).content
        path = f".//data//{pdf['name']}{pdf['title']}"+'.PDF'
        with open(path,'wb') as f:
            f.write(response)
        print(pdf['title']," 下载完毕",'*'*80)

def creat_ex(data):
    print('='*100,"\n写入表格",)
    global exclname
    all=[]
    href=[]
    if not os.path.exists(exclname):
        excel = ex.Workbook()
        sheet = excel.active
        sheet.append(['code','company','title','date'])
        excel.save(exclname)
    else:
        excel = ex.load_workbook(exclname)
        sheet = excel.active
        max_col = sheet.max_column
        sheet.delete_cols(idx=1,amount=max_col)
        sheet.append(['code','company','title','date'])
        excel.save(exclname)
    for it in data:
        ax_list = []
        hr = []
        ax_list.append(it['code'])
        ax_list.append(it['name'])
        ax_list.append(it['title'])
        ax_list.append(it['date'])
        hr.append(f'http://www.cninfo.com.cn/new/disclosure/stock?stockCode={it["code"]}&orgId={it["orgId"]}#latestAnnouncement')
        hr.append(f'http://www.cninfo.com.cn/new/disclosure/detail?plate=sse&orgId={it["orgId"]}&stockCode={it["code"]}&announcementId={it["announcementId"]}&announcementTime={it["date"]}')
        href.append(hr)
        all.append(ax_list)

    workbook = ex.load_workbook(exclname)
    sheet = workbook.active
    for row,h in zip(all,href):
        sheet.append(row)
        r = sheet.max_row
        sheet.cell(r,1).hyperlink=(h[0])
        sheet.cell(r,2).hyperlink=(h[0])
        sheet.cell(r,3).hyperlink =(h[1])
    workbook.save(exclname)
    print('写入完成')
    print('='*100,'\n')




if __name__ == '__main__' :
    '''
    key=关键字  
    page——number：爬取的页数  
    date：指定时间段  
    start：开始的页数  
    year：Ture表示年报  Flse表示不勾选
    colum: "szse" >深沪京 "hke" > 港股  "third" > 三板  "fund" >基金  ""
    巨潮网：http://www.cninfo.com.cn/new/commonUrl/pageOfSearch?url=disclosure/list/search&lastPage=index
    '''
    exclname = '1.xlsx'
    #拿到数据
    totle_new = get_data(key='',page_number=200,date="2022-01-01~2022-05-07",start=1,year=True,colum='szse')
    #简单的过滤数据
    useful_new=fenxidata(totle_new)
    #写入表格
    creat_ex(useful_new)
    #下载文件
    b=input("是否下载文件（y/n）>>")
    if b == 'y' :
        down_pdf(useful_new)
    print("运行完毕")
